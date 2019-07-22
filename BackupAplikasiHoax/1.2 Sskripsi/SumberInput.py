import Nlp
import openpyxl



#Proses Input data Uji

def Sumber(text):
    g=Nlp.Neuralp(text)
    return g

#uji Inputan
#text='Halo Susu dunia.. Hai('
#Hasilnlp=Sumber(text)
#print(Hasilnlp)

def hitung(text):

    #pembandingan dengan librari
    nlpcount=0
    text=text.split()
    jmlkata=len(text)
    sigmakatasama=0
    #print(jmlkata)
    for g in text:
        wk = openpyxl.load_workbook("kata.xlsx") #ganti
        sh = wk.active
        rows = sh.max_row
        for i in range(1, rows + 1):
                c = sh.cell(i,1)
                h=sh.cell(i,4)
                #print(c.value)
                if(text[nlpcount]==c.value):
                    sigmakatasama+=h.value
                   # print(c.value,h.value,Hasilnlp[nlpcount])
        nlpcount+=1
    print(sigmakatasama,jmlkata)
    ratahasil=sigmakatasama/jmlkata
    return ratahasil
#print(hitung(Hasilnlp))