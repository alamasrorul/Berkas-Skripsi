import pymysql
import  math
import openpyxl

#pembuatan librari kata
def library ():
    count=0
    loading=0;
    Kata=[]

    # Open database connection
    db = pymysql.connect("localhost", "xiinlaw", "Arcasevenvold04", "skripsi")

    # prepare a cursor object using cursor() method
    cursor = db.cursor()

    # Prepare SQL query to INSERT a record into the database.
    sql = "SELECT * FROM berita " #ganti
    try:
        # Execute the SQL command
        cursor.execute(sql)
        # Fetch all the rows in a list of lists.
        results = cursor.fetchall()
        roww = 2
        for row in results:


            berita = row[2]

            #tokenizing
            berita=berita.split()

           #proses pencarian atribut

            r = 0;
            loading+=1
            print((loading/50)*100," %")


            #perulangan indexing kata
            for g in berita:
                 print('Jumlah perulangan :',count)
                 count+=1

                 if (berita[r] in Kata):
                     print('skip')
                     r+=1
                 else :

                   Tf = berita.count(berita[r]) / len(berita)

                   # idf dan apriori

                   #jumlah dokumen yang mengandung kata
                   sqli = "SELECT ID FROM berita WHERE Berita LIKE %s" #ganti

                   b = berita[r]
                   a = ('%' + b + '%')
                   val = (a)

                   cursor.execute(sqli, val)
                   a = cursor.fetchall()
                   rr = len(a)
                   rr=float(rr)
                   Kata.append(berita[r])
  #                 print('Kata ke : ',r)

 #                  print(berita[r],' Jmlh :',rr)
                   Idf = math.log(100 / rr)
                   apriori = rr/100
                   TfIdf=Tf*Idf
#                   print('IDF : ',Idf,' apriori : ',apriori,'TF IDF : ',TfIdf)


                   # Save on Excel
                   wk = openpyxl.load_workbook("kata.xlsx") #ganti
                   sh=wk['kata']




                   # input kata on excel
                   kata = sh.cell(column=1, row=roww)
                   kata.value = berita[r]


                   # Weight
                   weight = sh.cell(column=2, row=roww)
                   weight.value = TfIdf


                   # Support
                   supp = sh.cell(column=3, row=roww)
                   supp.value = apriori

                   # Hasil
                   hasil = sh.cell(column=4, row=roww)
                   hasil.value = (apriori+TfIdf)/2

                   print('Jumlah Kata Yang tersimpan : ', roww-1)
                   r += 1
                   roww += 1


                   wk.save('uji.xlsx') #ganti



    except:
        print("Error: unable to fetch data")

    # menutup koneksi ke server

    db.close()

#run Library
library()

